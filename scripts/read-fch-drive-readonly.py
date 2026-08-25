#!/usr/bin/env python3
"""Leitura SOMENTE LEITURA dos FCH no Google Drive e normalização para o PMO.

Nunca faz upload, update, rename, delete, trash ou escrita no Google Drive.
Escopo OAuth: https://www.googleapis.com/auth/drive.readonly

Fontes aceitas:
- XLSX armazenado no Drive;
- Google Sheets nativo (exportado em memória como XLSX, sem alterar a origem).

Regra OPR/Madri:
- projeto contendo OPR e Madri => 100% para OPR + 100% para Madri;
- Madri exclusivo => Madri;
- OPR exclusivo => OPR;
- capacidade global deve usar a entrada de origem uma única vez (a duplicação é analítica).
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable

import requests
from google.auth.transport.requests import AuthorizedSession
from google.oauth2 import service_account
from google.oauth2.credentials import Credentials
from openpyxl import load_workbook

DRIVE_READONLY = "https://www.googleapis.com/auth/drive.readonly"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
GSHEET_MIME = "application/vnd.google-apps.spreadsheet"
FCH_PREFIX = "FCH - Formulário de Controle de Horas_"
CURRENT_FILE_ID = "1EwBf-iJyXsIKu5UoPP6iv-T9gnQQjPfH"

PEOPLE = [
    ("Victor Hugo", ["FCH - Victor Hugo", "FCH- Victor Hugo", "FCH Victor Hugo"]),
    ("Gabriel", ["FCH - Gabriel", "FCH-Gabriel", "FCH Gabriel"]),
]


def norm(value: Any) -> str:
    s = str(value or "").lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


def compact(value: Any) -> str:
    return norm(value).replace(" ", "")


def google_session() -> AuthorizedSession:
    raw_sa = os.getenv("GOOGLE_DRIVE_SERVICE_ACCOUNT_JSON", "").strip()
    if raw_sa:
        try:
            info = json.loads(raw_sa)
        except json.JSONDecodeError as exc:
            raise RuntimeError("GOOGLE_DRIVE_SERVICE_ACCOUNT_JSON não contém JSON válido") from exc
        creds = service_account.Credentials.from_service_account_info(info, scopes=[DRIVE_READONLY])
        return AuthorizedSession(creds)

    client_id = os.getenv("GOOGLE_CLIENT_ID", "").strip()
    client_secret = os.getenv("GOOGLE_CLIENT_SECRET", "").strip()
    refresh_token = os.getenv("GOOGLE_REFRESH_TOKEN", "").strip()
    if client_id and client_secret and refresh_token:
        creds = Credentials(
            token=None,
            refresh_token=refresh_token,
            token_uri="https://oauth2.googleapis.com/token",
            client_id=client_id,
            client_secret=client_secret,
            scopes=[DRIVE_READONLY],
        )
        return AuthorizedSession(creds)

    raise RuntimeError(
        "Credenciais Google read-only ausentes. Configure GOOGLE_DRIVE_SERVICE_ACCOUNT_JSON "
        "ou GOOGLE_CLIENT_ID + GOOGLE_CLIENT_SECRET + GOOGLE_REFRESH_TOKEN."
    )


def drive_get_json(session: AuthorizedSession, url: str, *, params: dict[str, str] | None = None) -> dict[str, Any]:
    r = session.get(url, params=params, timeout=60)
    if not r.ok:
        raise RuntimeError(f"Google Drive HTTP {r.status_code}: {r.text[:600]}")
    return r.json()


def list_sources(session: AuthorizedSession, file_id: str, folder_id: str) -> list[dict[str, Any]]:
    fields = "id,name,mimeType,modifiedTime,size,parents"
    if folder_id:
        q = (
            f"'{folder_id}' in parents and trashed=false and "
            f"name contains '{FCH_PREFIX.replace("'", "\\'")}'"
        )
        files: list[dict[str, Any]] = []
        token = ""
        while True:
            params = {
                "q": q,
                "fields": f"nextPageToken,files({fields})",
                "pageSize": "100",
                "orderBy": "name",
                "supportsAllDrives": "true",
                "includeItemsFromAllDrives": "true",
            }
            if token:
                params["pageToken"] = token
            data = drive_get_json(session, "https://www.googleapis.com/drive/v3/files", params=params)
            files.extend(data.get("files") or [])
            token = str(data.get("nextPageToken") or "")
            if not token:
                break
        return [f for f in files if f.get("mimeType") in {XLSX_MIME, GSHEET_MIME}]

    # Arquivo corrente como fallback seguro. Nunca procura/edita arquivos fora do escopo indicado.
    fid = file_id or CURRENT_FILE_ID
    data = drive_get_json(
        session,
        f"https://www.googleapis.com/drive/v3/files/{fid}",
        params={"fields": fields, "supportsAllDrives": "true"},
    )
    return [data]


def download_xlsx(session: AuthorizedSession, meta: dict[str, Any]) -> bytes:
    fid = str(meta["id"])
    mime = str(meta.get("mimeType") or "")
    if mime == XLSX_MIME:
        url = f"https://www.googleapis.com/drive/v3/files/{fid}"
        r = session.get(url, params={"alt": "media", "supportsAllDrives": "true"}, timeout=120)
    elif mime == GSHEET_MIME:
        url = f"https://www.googleapis.com/drive/v3/files/{fid}/export"
        r = session.get(url, params={"mimeType": XLSX_MIME}, timeout=120)
    else:
        raise RuntimeError(f"Formato FCH não suportado: {mime} ({meta.get('name')})")
    if not r.ok:
        raise RuntimeError(f"Falha ao baixar {meta.get('name')}: HTTP {r.status_code} {r.text[:300]}")
    return bytes(r.content)


def duration_hours(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, timedelta):
        return value.total_seconds() / 3600.0
    if isinstance(value, datetime):
        return value.hour + value.minute / 60.0 + value.second / 3600.0
    if isinstance(value, time):
        return value.hour + value.minute / 60.0 + value.second / 3600.0
    if isinstance(value, (int, float)):
        n = float(value)
        if n <= 0:
            return 0.0
        return n * 24.0 if n <= 1.5 else n
    s = str(value).strip()
    m = re.fullmatch(r"(\d+):([0-5]\d)(?::([0-5]\d))?", s)
    if m:
        return int(m.group(1)) + int(m.group(2)) / 60.0 + int(m.group(3) or 0) / 3600.0
    try:
        n = float(s.replace(" ", "").replace(",", "."))
    except ValueError:
        return 0.0
    return n * 24.0 if 0 < n <= 1.5 else max(0.0, n)


def iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def find_sheet(wb, aliases: Iterable[str]):
    by_norm = {norm(ws.title): ws for ws in wb.worksheets}
    for alias in aliases:
        if norm(alias) in by_norm:
            return by_norm[norm(alias)]
    return None


def find_header(ws) -> tuple[int, int, int, int] | None:
    for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True), start=1):
        vals = [norm(v) for v in row]
        project = next((i for i, v in enumerate(vals) if v == "projeto"), -1)
        hours = next((i for i, v in enumerate(vals) if "tempo da atividade" in v), -1)
        dt = next((i for i, v in enumerate(vals) if v.startswith("data")), -1)
        if project >= 0 and hours >= 0 and dt >= 0:
            return row_no, dt, hours, project
    return None


def map_targets(project: str) -> list[tuple[str, str, str]]:
    p = compact(project)
    if not p:
        return []
    if "opr" in p and ("madri" in p or "madrid" in p):
        return [
            ("OPR", "OPR", "OPR_Madri compartilhado"),
            ("Madri", "Madri", "OPR_Madri compartilhado"),
        ]
    if "madri" in p or "madrid" in p:
        return [("Madri", "Madri", "Madri exclusivo")]
    if p == "opr" or "rfpopr" in p or "pmoopr" in p:
        return [("OPR", "OPR", "OPR exclusivo")]
    return []


@dataclass(frozen=True)
class CsvRow:
    empresa: str
    projeto: str
    hora: float
    mes: str
    consultor: str
    projeto_origem: str
    origem_compartilhada: str


def parse_file(meta: dict[str, Any], content: bytes) -> list[CsvRow]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out: list[CsvRow] = []
    try:
        for person, aliases in PEOPLE:
            ws = find_sheet(wb, aliases)
            if ws is None:
                print(f"[fch] aviso: aba de {person} não encontrada em {meta.get('name')}", file=sys.stderr)
                continue
            hdr = find_header(ws)
            if hdr is None:
                print(f"[fch] aviso: cabeçalho não encontrado em {ws.title}", file=sys.stderr)
                continue
            header_row, date_idx, hours_idx, project_idx = hdr
            for row_no, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                if max(date_idx, hours_idx, project_idx) >= len(row):
                    continue
                project = str(row[project_idx] or "").strip()
                hours = duration_hours(row[hours_idx])
                activity_date = iso_date(row[date_idx])
                if not project or hours <= 0 or not activity_date:
                    continue
                month = activity_date[:7]
                targets = map_targets(project)
                shared = "sim" if len(targets) == 2 else "nao"
                for company, target_project, _rule in targets:
                    out.append(
                        CsvRow(
                            empresa=company,
                            projeto=target_project,
                            hora=round(hours, 4),
                            mes=month,
                            consultor=person,
                            projeto_origem=project,
                            origem_compartilhada=shared,
                        )
                    )
    finally:
        wb.close()
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", required=True)
    ap.add_argument("--file-id", default=os.getenv("FCH_FILE_ID", CURRENT_FILE_ID))
    ap.add_argument("--folder-id", default=os.getenv("FCH_FOLDER_ID", ""))
    args = ap.parse_args()

    session = google_session()
    sources = list_sources(session, args.file_id.strip(), args.folder_id.strip())
    if not sources:
        raise RuntimeError("Nenhum FCH encontrado no escopo read-only configurado")

    rows: list[CsvRow] = []
    for meta in sources:
        name = str(meta.get("name") or "")
        if args.folder_id and not name.startswith(FCH_PREFIX):
            continue
        print(f"[fch] lendo somente: {name} ({meta.get('id')})")
        rows.extend(parse_file(meta, download_xlsx(session, meta)))

    if not rows:
        raise RuntimeError("Nenhuma hora OPR/Madri encontrada nos FCH lidos")

    path = Path(args.output)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["empresa", "projeto", "hora", "mes", "consultor", "projeto_origem", "origem_compartilhada"])
        for r in rows:
            writer.writerow([r.empresa, r.projeto, f"{r.hora:.4f}", r.mes, r.consultor, r.projeto_origem, r.origem_compartilhada])

    totals: dict[str, float] = {}
    for r in rows:
        totals[r.empresa] = totals.get(r.empresa, 0.0) + r.hora
    print("[fch] alocações:", len(rows))
    print("[fch] totais analíticos:", json.dumps({k: round(v, 4) for k, v in totals.items()}, ensure_ascii=False))
    print(f"[fch] CSV temporário: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
