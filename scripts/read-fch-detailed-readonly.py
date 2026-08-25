#!/usr/bin/env python3
"""Extrai FCH do Google Drive em modo estritamente read-only para a Curva S."""
from __future__ import annotations
import argparse, io, json, os, re, unicodedata, hashlib
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any
from google.auth.transport.requests import AuthorizedSession
from google.oauth2 import service_account
from google.oauth2.credentials import Credentials
from openpyxl import load_workbook

SCOPE='https://www.googleapis.com/auth/drive.readonly'
XLSX='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
GSHEET='application/vnd.google-apps.spreadsheet'
DEFAULT_FILE='1EwBf-iJyXsIKu5UoPP6iv-T9gnQQjPfH'
PREFIX='FCH - Formulário de Controle de Horas_'
PEOPLE=[('Victor Hugo',['FCH - Victor Hugo','FCH- Victor Hugo','FCH Victor Hugo']),('Gabriel',['FCH - Gabriel','FCH-Gabriel','FCH Gabriel'])]

def norm(v:Any)->str:
 s=unicodedata.normalize('NFD',str(v or '').lower());s=''.join(c for c in s if unicodedata.category(c)!='Mn');return re.sub(r'[^a-z0-9]+',' ',s).strip()
def compact(v:Any)->str:return norm(v).replace(' ','')
def session():
 raw=os.getenv('GOOGLE_DRIVE_SERVICE_ACCOUNT_JSON','').strip()
 if raw:return AuthorizedSession(service_account.Credentials.from_service_account_info(json.loads(raw),scopes=[SCOPE]))
 cid=os.getenv('GOOGLE_CLIENT_ID','').strip();sec=os.getenv('GOOGLE_CLIENT_SECRET','').strip();ref=os.getenv('GOOGLE_REFRESH_TOKEN','').strip()
 if cid and sec and ref:return AuthorizedSession(Credentials(token=None,refresh_token=ref,token_uri='https://oauth2.googleapis.com/token',client_id=cid,client_secret=sec,scopes=[SCOPE]))
 raise RuntimeError('Credencial Google read-only ausente')
def getj(s,url,params=None):
 r=s.get(url,params=params or {},timeout=60)
 if not r.ok:raise RuntimeError(f'Google Drive HTTP {r.status_code}: {r.text[:300]}')
 return r.json()
def sources(s,file_id,folder_id):
 fields='id,name,mimeType,modifiedTime,size'
 if not folder_id:return [getj(s,f'https://www.googleapis.com/drive/v3/files/{file_id or DEFAULT_FILE}',{'fields':fields,'supportsAllDrives':'true'})]
 out=[];token=''
 while True:
  p={'q':f"'{folder_id}' in parents and trashed=false and name contains 'FCH - Formulário de Controle de Horas_'",'fields':f'nextPageToken,files({fields})','pageSize':'100','orderBy':'name','supportsAllDrives':'true','includeItemsFromAllDrives':'true'}
  if token:p['pageToken']=token
  d=getj(s,'https://www.googleapis.com/drive/v3/files',p);out.extend(d.get('files') or []);token=str(d.get('nextPageToken') or '')
  if not token:break
 return [x for x in out if x.get('mimeType') in {XLSX,GSHEET}]
def download(s,m):
 fid=m['id'];mime=m.get('mimeType')
 if mime==XLSX:r=s.get(f'https://www.googleapis.com/drive/v3/files/{fid}',params={'alt':'media','supportsAllDrives':'true'},timeout=120)
 elif mime==GSHEET:r=s.get(f'https://www.googleapis.com/drive/v3/files/{fid}/export',params={'mimeType':XLSX},timeout=120)
 else:raise RuntimeError('Formato não suportado')
 if not r.ok:raise RuntimeError(f'Download falhou: HTTP {r.status_code}')
 return bytes(r.content)
def hours(v):
 if v in (None,''):return 0.0
 if isinstance(v,timedelta):return v.total_seconds()/3600
 if isinstance(v,datetime):return v.hour+v.minute/60+v.second/3600
 if isinstance(v,time):return v.hour+v.minute/60+v.second/3600
 if isinstance(v,(int,float)):
  n=float(v);return n*24 if 0<n<=1.5 else max(0,n)
 m=re.fullmatch(r'(\d+):([0-5]\d)(?::([0-5]\d))?',str(v).strip())
 if m:return int(m.group(1))+int(m.group(2))/60+int(m.group(3) or 0)/3600
 try:
  n=float(str(v).strip().replace(',','.'));return n*24 if 0<n<=1.5 else max(0,n)
 except:return 0.0
def isod(v):
 if isinstance(v,datetime):return v.date().isoformat()
 if isinstance(v,date):return v.isoformat()
 s=str(v or '').strip()
 for f in ('%d/%m/%Y','%Y-%m-%d','%Y-%m-%d %H:%M:%S'):
  try:return datetime.strptime(s[:19],f).date().isoformat()
  except:pass
 return ''
def target(project):
 p=compact(project)
 if 'opr' in p and ('madri' in p or 'madrid' in p):return [('OPR','OPR_Madri compartilhado'),('MADRI','OPR_Madri compartilhado')]
 if 'madri' in p or 'madrid' in p:return [('MADRI','Madri exclusivo')]
 if p=='opr' or 'rfpopr' in p or 'pmoopr' in p:return [('OPR','OPR exclusivo')]
 return []
def header(ws):
 for rn,row in enumerate(ws.iter_rows(min_row=1,max_row=min(ws.max_row,20),values_only=True),1):
  vals=[norm(x) for x in row];pi=next((i for i,v in enumerate(vals) if v=='projeto'),-1);hi=next((i for i,v in enumerate(vals) if 'tempo da atividade' in v or v=='horas'),-1);di=next((i for i,v in enumerate(vals) if v.startswith('data')),-1)
  if min(pi,hi,di)>=0:return rn,di,hi,pi
 return None
def parse(meta,content):
 wb=load_workbook(io.BytesIO(content),read_only=True,data_only=True);out=[]
 try:
  by={norm(w.title):w for w in wb.worksheets}
  for person,aliases in PEOPLE:
   ws=next((by[norm(a)] for a in aliases if norm(a) in by),None)
   if not ws:continue
   h=header(ws)
   if not h:continue
   hr,di,hi,pi=h
   for row_no,row in enumerate(ws.iter_rows(min_row=hr+1,values_only=True),hr+1):
    if max(di,hi,pi)>=len(row):continue
    proj=str(row[pi] or '').strip();hrs=hours(row[hi]);dt=isod(row[di])
    if not proj or hrs<=0 or not dt:continue
    tg=target(proj)
    if not tg:continue
    raw='|'.join([str(meta['id']),ws.title,str(row_no),dt,person,proj,f'{hrs:.4f}']);hashv=hashlib.sha256(raw.encode()).hexdigest()
    for t,rule in tg:out.append({'source_file_id':str(meta['id']),'source_file_name':str(meta.get('name') or ''),'source_modified_at':str(meta.get('modifiedTime') or ''),'source_sheet':ws.title,'source_row':row_no,'person':person,'activity_date':dt,'source_project':proj,'target_project':t,'allocation_rule':rule,'source_entry_hash':hashv,'hours':round(hrs,4)})
 finally:wb.close()
 return out

def main():
 ap=argparse.ArgumentParser();ap.add_argument('--output',required=True);ap.add_argument('--file-id',default=os.getenv('FCH_FILE_ID',DEFAULT_FILE));ap.add_argument('--folder-id',default=os.getenv('FCH_FOLDER_ID',''));a=ap.parse_args();s=session();src=sources(s,a.file_id.strip(),a.folder_id.strip());entries=[]
 for m in src:
  if a.folder_id and not str(m.get('name') or '').startswith(PREFIX):continue
  print('[fch-detail] lendo somente:',m.get('name'),m.get('id'));entries.extend(parse(m,download(s,m)))
 if not entries:raise RuntimeError('Nenhuma hora OPR/MADRI encontrada')
 unique={e['source_entry_hash']:e['hours'] for e in entries};tot={'OPR':0.0,'MADRI':0.0}
 for e in entries:tot[e['target_project']]+=e['hours']
 payload={'policy':'google-drive-readonly','sources':[{'id':m.get('id'),'name':m.get('name'),'modified_at':m.get('modifiedTime')} for m in src],'entries':entries,'summary':{'allocations':len(entries),'source_entries':len(unique),'capacity_hours':round(sum(unique.values()),4),'opr_hours':round(tot['OPR'],4),'madri_hours':round(tot['MADRI'],4)}}
 Path(a.output).write_text(json.dumps(payload,ensure_ascii=False),encoding='utf-8');print('[fch-detail]',json.dumps(payload['summary'],ensure_ascii=False));return 0
if __name__=='__main__':raise SystemExit(main())
